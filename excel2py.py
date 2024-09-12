import os

from openpyxl import load_workbook

# 加载 Excel 文件
xml_path = input("请输入 XML 文件绝对路径：")

# 对路径进行判断
if not os.path.exists(xml_path):
    print("文件不存在")
    exit()

workbook = load_workbook(xml_path)

sheet_names = workbook.get_sheet_names()

test_case_sheet = workbook.get_sheet_by_name("测试用例")


def get_test_case_info():
    """获取测试用例信息"""
    test_case_dict = {}
    for row in test_case_sheet.iter_rows(min_row=2, values_only=True):
        test_case_dict[row[0]] = {
            "测试用例编号": row[0],
            "对应需求编号": row[1],
            "用例名称": row[2],
            "用例描述": row[3],
            "标签": row[4],
            "优先级": row[5],
            "前置条件": row[6],
            "用例测试步骤": row[7],
            "预期结果": row[8],
        }
    return test_case_dict


def generate_test_case(test_case: dict, suite_name: str):
    """
    生成测试用例
    :param test_case: 测试用例信息
    :param suite_name: 套件名称
    """
    # 判断字典是否为空
    if not test_case:
        print("test_case is empty")
        exit()
    # 判断suit_name是否符合英文且无其他符号
    if suite_name and not suite_name.isalpha():
        print("suite_name must be english and no other symbols")
        exit()

    # 生成测试用例
    output_file = f"test_{suite_name}.py"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("import allure\n")
        f.write("import pytest\n\n")

        f.write(f"@allure.suite('{suite_name}')\n")
        f.write(f"class Test{suite_name}:\n")

        # 遍历测试用例字典
        for test_case_name, test_case_info in test_case.items():
            f.write(f"    @allure.title('{test_case_info['用例名称']}')\n")
            f.write(f"    @allure.description('{test_case_info['用例描述']}')\n")
            f.write(f"    def test_{test_case_name}(self):\n")
            f.write(f"        \"\"\"\n")

            f.write(f"        前置条件：\n")
            # 循环打印前置条件
            for line in test_case_info['前置条件'].splitlines():
                f.write(f"        {line}\n")

            f.write(f"        用例测试步骤：\n")
            # 循环用例测试步骤
            for line in test_case_info['用例测试步骤'].splitlines():
                f.write(f"        {line}\n")

            f.write(f"        预期结果：\n")
            # 循环打印预期结果
            for line in test_case_info['预期结果'].splitlines():
                f.write(f"        {line}\n")
            f.write(f"        \"\"\"\n")
            for line in test_case_info['用例测试步骤'].splitlines():
                f.write(f"        with allure.step(\"{line}\"):\n")
                f.write(f"            assert True\n\n")


if __name__ == "__main__":
    # 生成测试用例
    test_case = get_test_case_info()
    suite_name = input("请输入测试套件名称（英文）：")
    generate_test_case(test_case, suite_name)
